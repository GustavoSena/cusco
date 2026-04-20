"""PRR (Plano de Recuperação e Resiliência) data source.

Downloads two bulk XLSX datasets from dados.gov.pt:

- PRR Entidades (~750K rows): funding records per entity, tracked by NIF. Each
  row records a project/entity pairing with contracted and paid values, role
  (beneficiário final, intermediário, etc.), CAE code and municipality.
- PRR Contratos (~12K rows): public contracts signed under PRR funding, with
  contracting authority / supplier roles.

Both are loaded once on a 24h TTL and persisted in SQLite
(:mod:`cusco.storage`) — the DB IS the cache now, so warm restarts don't
re-download or re-parse.
"""

from __future__ import annotations

import asyncio
import datetime as _dt
import io
import logging
from typing import Any

from ..models import PRRContract, PRRFunding
from ..storage import BulkTable
from .base import DataSource

logger = logging.getLogger(__name__)

DATASET_API_ENTIDADES = (
    "https://dados.gov.pt/api/1/datasets/"
    "dataset-estrutura-de-missao-prr-entidades-1/"
)
DATASET_API_CONTRATOS = (
    "https://dados.gov.pt/api/1/datasets/"
    "dataset-estrutura-de-missao-prr-entidades-contratos-publicos/"
)

CACHE_MAX_AGE_SECONDS = 24 * 3600  # 1 day


def _safe_float(val: Any) -> float | None:
    if val is None or val == "":
        return None
    try:
        return float(str(val).replace(",", ".").replace(" ", ""))
    except (ValueError, TypeError):
        return None


def _json_safe(value: Any) -> Any:
    """Convert openpyxl cell values to JSON-serializable equivalents.
    Datetime and date cells become ISO strings; everything else passes through.
    """
    if isinstance(value, (_dt.datetime, _dt.date)):
        return value.isoformat()
    return value


def _normalize_nif(val: Any) -> str:
    """Normalize NIF-like values: strip, remove PT prefix, keep digits as-is."""
    if val is None:
        return ""
    s = str(val).strip()
    if s.upper().startswith("PT"):
        s = s[2:].strip()
    # Some NIFs come as floats (openpyxl quirk): "514181435.0"
    if s.endswith(".0"):
        s = s[:-2]
    return s


class PRRSource(DataSource):
    """PRR entity funding + contracts from dados.gov.pt."""

    name = "prr"

    def __init__(self, timeout: float = 120.0):
        super().__init__(timeout=timeout)
        # Fundings keyed by nif_entidade, contracts keyed by cd_entidade
        # (which is also a NIF-like value).
        self._fundings_table = BulkTable("prr_fundings")
        self._contracts_table = BulkTable("prr_contracts")
        self._loaded = False

    async def _ensure_loaded(self) -> None:
        """Idempotent load — serialized via `_load_once` so concurrent
        first queries don't each re-download + re-parse the dataset."""
        await self._load_once(self._do_load, lambda: self._loaded)

    async def _do_load(self) -> None:
        fundings_fresh = await asyncio.to_thread(
            self._fundings_table.is_fresh, CACHE_MAX_AGE_SECONDS
        )
        contracts_fresh = await asyncio.to_thread(
            self._contracts_table.is_fresh, CACHE_MAX_AGE_SECONDS
        )
        if fundings_fresh and contracts_fresh:
            f_count = await asyncio.to_thread(self._fundings_table.row_count)
            c_count = await asyncio.to_thread(self._contracts_table.row_count)
            logger.info(
                f"PRR: SQLite fresh — {f_count} funding rows, "
                f"{c_count} contract rows (skipping download)"
            )
            self._loaded = True
            return

        # Track whether either side refused the new payload (empty
        # download, parse failure with 0 usable rows) — don't flip
        # `_loaded=True` if we failed to refresh a stale table, so the
        # next query gets another shot.
        fundings_refreshed = fundings_fresh
        contracts_refreshed = contracts_fresh

        if not fundings_fresh:
            try:
                fundings = await self._load_fundings()
                rows = [
                    (nif, row)
                    for row in fundings
                    if (nif := _normalize_nif(row.get("nif_entidade")))
                ]
                if rows:
                    await asyncio.to_thread(
                        self._fundings_table.replace_all, rows
                    )
                    fundings_refreshed = True
                else:
                    # Keep yesterday's cache rather than wiping to zero
                    # when upstream went AWOL.
                    logger.warning(
                        "PRR entidades returned 0 usable rows — "
                        "keeping previous cache"
                    )
            except Exception as e:  # noqa: BLE001
                logger.warning(f"Failed to load PRR entidades: {e}")

        if not contracts_fresh:
            try:
                contracts = await self._load_contracts()
                rows = [
                    (nif, row)
                    for row in contracts
                    if (nif := _normalize_nif(row.get("cd_entidade")))
                ]
                if rows:
                    await asyncio.to_thread(
                        self._contracts_table.replace_all, rows
                    )
                    contracts_refreshed = True
                else:
                    logger.warning(
                        "PRR contratos returned 0 usable rows — "
                        "keeping previous cache"
                    )
            except Exception as e:  # noqa: BLE001
                logger.warning(f"Failed to load PRR contratos: {e}")

        if fundings_refreshed and contracts_refreshed:
            self._loaded = True
            f_count = await asyncio.to_thread(self._fundings_table.row_count)
            c_count = await asyncio.to_thread(self._contracts_table.row_count)
            logger.info(
                f"Indexed PRR: {f_count} funding rows, {c_count} contract rows"
            )
        else:
            logger.warning(
                f"PRR load incomplete (fundings={fundings_refreshed}, "
                f"contracts={contracts_refreshed}) — will retry on next query"
            )

    # -- Dataset A: Entidades ----------------------------------------------
    async def _load_fundings(self) -> list[dict]:
        url = await self._find_resource_url(DATASET_API_ENTIDADES, "entidades")
        if not url:
            logger.warning("Could not find PRR entidades download URL")
            return []

        async with self._client() as client:
            logger.info(f"Downloading PRR entidades from {url}...")
            resp = await client.get(url)
            resp.raise_for_status()
            rows = self._parse_prr_entidades_xlsx(resp.content)
            logger.info(f"Loaded {len(rows)} PRR entidades rows")
            return rows

    # Columns we actually read downstream — parser drops everything else to
    # minimize row size (750K rows × unused columns = ~hundreds of MB).
    _ENTIDADES_COLUMNS = frozenset([
        "nif_entidade",
        "ds_entidade",
        "papel_entidade",
        "atividade_economica",
        "localizacao_sede",
        "valor_contratado",
        "valor_pago",
        "dt_referencia",
        "cd_projeto",
    ])

    _CONTRATOS_COLUMNS = frozenset([
        "cd_entidade",
        "cd_contrato",
        "ds_contrato",
        "ds_entidade",
        "ds_papel_entidade_contrato",
        "valor_contrato",
        "dt_referencia",
    ])

    def _parse_xlsx_rows(
        self, xlsx_bytes: bytes, keep_columns: frozenset[str]
    ) -> list[dict]:
        """Parse an XLSX workbook into list[dict], retaining only the columns
        listed in ``keep_columns``. Unknown/empty headers are dropped entirely,
        which keeps the persisted payload small even for 750K+ row datasets."""
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
        try:
            ws = wb.active
            if ws is None:
                return []

            rows_iter = ws.iter_rows(values_only=True)
            try:
                headers = [
                    str(h).strip() if h is not None else "" for h in next(rows_iter)
                ]
            except StopIteration:
                return []

            # Case-insensitive header matching. If dados.gov.pt ever
            # publishes a revision with `NIF_Entidade` (vs. the current
            # all-lowercase `nif_entidade`), a strict comparison would
            # silently drop every column — every row would then look
            # "empty", we'd persist 0 rows, and the log would look
            # perfectly healthy. Normalize on both sides and preserve
            # the canonical (frozenset) casing as the output key.
            keep_lookup = {c.lower(): c for c in keep_columns}
            keep_idx = [
                (idx, keep_lookup[header.lower()])
                for idx, header in enumerate(headers)
                if header.lower() in keep_lookup
            ]
            if not keep_idx:
                logger.warning(
                    f"PRR XLSX headers did not match any expected columns "
                    f"(got {headers[:8]}…); persisted 0 rows"
                )
                return []

            out: list[dict] = []
            for row in rows_iter:
                if row is None:
                    continue
                record = {}
                for idx, header in keep_idx:
                    raw_val = row[idx] if idx < len(row) else None
                    record[header] = _json_safe(raw_val)
                if any(v not in (None, "") for v in record.values()):
                    out.append(record)
            return out
        finally:
            # `wb.close()` used to be on the happy path only — a parse
            # error above would leak the open workbook's file handle.
            wb.close()

    def _parse_prr_entidades_xlsx(self, xlsx_bytes: bytes) -> list[dict]:
        return self._parse_xlsx_rows(xlsx_bytes, self._ENTIDADES_COLUMNS)

    def _to_funding(self, raw: dict) -> PRRFunding:
        return PRRFunding(
            project_code=str(raw.get("cd_projeto") or "").strip(),
            entity_name=str(raw.get("ds_entidade") or "").strip(),
            role=str(raw.get("papel_entidade") or "").strip(),
            cae_code=str(raw.get("atividade_economica") or "").strip(),
            municipality=str(raw.get("localizacao_sede") or "").strip(),
            value_contracted=_safe_float(raw.get("valor_contratado")),
            value_paid=_safe_float(raw.get("valor_pago")),
            reference_date=str(raw.get("dt_referencia") or "").strip(),
        )

    # -- Dataset B: Contratos ----------------------------------------------
    async def _load_contracts(self) -> list[dict]:
        url = await self._find_resource_url(DATASET_API_CONTRATOS, "contratos")
        if not url:
            logger.warning("Could not find PRR contratos download URL")
            return []

        async with self._client() as client:
            logger.info(f"Downloading PRR contratos from {url}...")
            resp = await client.get(url)
            resp.raise_for_status()
            rows = self._parse_prr_contratos_xlsx(resp.content)
            logger.info(f"Loaded {len(rows)} PRR contratos rows")
            return rows

    def _parse_prr_contratos_xlsx(self, xlsx_bytes: bytes) -> list[dict]:
        return self._parse_xlsx_rows(xlsx_bytes, self._CONTRATOS_COLUMNS)

    def _to_contract(self, raw: dict) -> PRRContract:
        return PRRContract(
            contract_code=str(raw.get("cd_contrato") or "").strip(),
            description=str(raw.get("ds_contrato") or "").strip(),
            entity_name=str(raw.get("ds_entidade") or "").strip(),
            role=str(raw.get("ds_papel_entidade_contrato") or "").strip(),
            value=_safe_float(raw.get("valor_contrato")),
            reference_date=str(raw.get("dt_referencia") or "").strip(),
        )

    # -- Resource URL resolution -------------------------------------------
    async def _find_resource_url(self, dataset_api: str, hint: str) -> str | None:
        """Resolve the XLSX download URL from a dados.gov.pt dataset API."""
        try:
            async with self._client() as client:
                resp = await client.get(dataset_api)
                resp.raise_for_status()
                dataset = resp.json()

                hint_l = hint.lower()
                for resource in dataset.get("resources", []):
                    title = (resource.get("title") or "").lower()
                    url = resource.get("url") or ""
                    url_l = url.lower()
                    if not url:
                        continue
                    if url_l.endswith(".xlsx") and (
                        hint_l in title or hint_l in url_l
                    ):
                        return url
                # Fall back to the first xlsx resource if no hint match
                for resource in dataset.get("resources", []):
                    url = resource.get("url") or ""
                    if url.lower().endswith(".xlsx"):
                        return url
        except Exception as e:
            logger.warning(f"Failed to resolve PRR dataset URL ({dataset_api}): {e}")
        return None

    # -- Public API --------------------------------------------------------
    async def search_by_nif(self, nif: str) -> dict[str, Any]:
        await self._ensure_loaded()
        nif = _normalize_nif(nif)

        raw_fundings = await asyncio.to_thread(
            self._fundings_table.get_by_nif, nif
        )
        raw_contracts = await asyncio.to_thread(
            self._contracts_table.get_by_nif, nif
        )

        fundings = [self._to_funding(r) for r in raw_fundings]
        contracts = [self._to_contract(r) for r in raw_contracts]

        total_contracted = sum(f.value_contracted or 0 for f in fundings)
        total_paid = sum(f.value_paid or 0 for f in fundings)

        return {
            "prr_fundings": fundings,
            "prr_contracts": contracts,
            "has_prr_funding": bool(fundings),
            "prr_total_contracted": total_contracted,
            "prr_total_paid": total_paid,
        }
