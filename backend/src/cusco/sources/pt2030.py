"""Portugal 2030 programme funding source.

Downloads a single XLSX (~2MB, ~25K rows) from dados.gov.pt listing entities
benefiting from PT2030 structural fund operations, with contract values, fund
approved/executed/paid breakdown and the framework under which each operation
was financed.

Loaded once at startup, cached as JSON in `CUSCO_CACHE_DIR` (24h TTL), and
indexed by NIF for fast lookup.
"""

from __future__ import annotations

import datetime as _dt
import io
import json
import logging
import os
import time
from pathlib import Path
from typing import Any

from ..models import PT2030Funding
from .base import DataSource

logger = logging.getLogger(__name__)

DATASET_API = (
    "https://dados.gov.pt/api/1/datasets/"
    "datasets-pt2030-05-lista-de-entidades-pt2030/"
)

CACHE_DIR = Path(os.environ.get("CUSCO_CACHE_DIR", "/tmp/cusco_cache"))
CACHE_MAX_AGE_SECONDS = 24 * 3600  # 1 day


def _safe_float(val: Any) -> float | None:
    if val is None or val == "":
        return None
    try:
        return float(str(val).replace(",", ".").replace(" ", ""))
    except (ValueError, TypeError):
        return None


def _normalize_nif(val: Any) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    if s.upper().startswith("PT"):
        s = s[2:].strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


class PT2030Source(DataSource):
    """PT2030 entity funding from dados.gov.pt."""

    name = "pt2030"

    def __init__(self, timeout: float = 60.0):
        super().__init__(timeout=timeout)
        self._by_nif: dict[str, list[dict]] = {}
        self._loaded = False
        CACHE_DIR.mkdir(parents=True, exist_ok=True)

    async def _ensure_loaded(self) -> None:
        """Idempotent load — serialized via `_load_once` so concurrent
        first queries don't each re-download + re-parse the dataset."""
        await self._load_once(self._do_load, lambda: self._loaded)

    async def _do_load(self) -> None:
        try:
            rows = await self._load_rows()
            self._index(rows)
        except Exception as e:
            logger.warning(f"Failed to load PT2030 entidades: {e}")
        self._loaded = True
        logger.info(f"Indexed PT2030: {len(self._by_nif)} NIFs")
    def _read_cache(self, path: Path) -> list[dict] | None:
        """Read a cached JSON file. Returns None if missing, corrupt, or
        unreadable — caller should re-download in that case."""
        if not path.exists():
            return None
        try:
            with open(path) as f:
                return json.load(f)
        except (OSError, json.JSONDecodeError) as e:
            logger.warning(f"PT2030 cache {path.name} unreadable ({e}); re-downloading")
            try:
                path.unlink()
            except OSError:
                pass
            return None

    def _write_cache_atomic(self, path: Path, data: list[dict]) -> None:
        """Write JSON atomically (tmp file + rename) so a crash mid-write
        doesn't leave a corrupt cache."""
        tmp = path.with_suffix(path.suffix + ".tmp")
        try:
            with open(tmp, "w") as f:
                json.dump(data, f)
            os.replace(tmp, path)
        except Exception:
            try:
                tmp.unlink()
            except OSError:
                pass
            raise

    async def _load_rows(self) -> list[dict]:
        cache_file = CACHE_DIR / "pt2030_entidades.json"

        if cache_file.exists():
            age = time.time() - cache_file.stat().st_mtime
            if age < CACHE_MAX_AGE_SECONDS:
                logger.info("Loading PT2030 entidades from cache")
                cached = self._read_cache(cache_file)
                if cached is not None:
                    return cached

        url = await self._find_xlsx_url()
        if not url:
            logger.warning("Could not find PT2030 entidades download URL")
            return self._read_cache(cache_file) or []

        async with self._client() as client:
            logger.info(f"Downloading PT2030 entidades from {url}...")
            resp = await client.get(url)
            resp.raise_for_status()
            rows = self._parse_xlsx(resp.content)

            self._write_cache_atomic(cache_file, rows)
            logger.info(f"Loaded {len(rows)} PT2030 entidades rows")
            return rows

    def _parse_xlsx(self, xlsx_bytes: bytes) -> list[dict]:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            return []

        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
        except StopIteration:
            return []

        out: list[dict] = []
        for row in rows_iter:
            if row is None:
                continue
            record = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                value = row[idx] if idx < len(row) else None
                # openpyxl returns datetime objects for date cells — not JSON
                # serializable. Convert to ISO strings during parse so the
                # cached JSON round-trips cleanly.
                if isinstance(value, (_dt.datetime, _dt.date)):
                    value = value.isoformat()
                record[header] = value
            if any(v not in (None, "") for v in record.values()):
                out.append(record)

        wb.close()
        return out

    async def _find_xlsx_url(self) -> str | None:
        try:
            async with self._client() as client:
                resp = await client.get(DATASET_API)
                resp.raise_for_status()
                dataset = resp.json()

                for resource in dataset.get("resources", []):
                    url = resource.get("url") or ""
                    if url.lower().endswith(".xlsx"):
                        return url
        except Exception as e:
            logger.warning(f"Failed to resolve PT2030 dataset URL: {e}")
        return None

    def _index(self, rows: list[dict]) -> None:
        for row in rows:
            nif = _normalize_nif(row.get("Nif entidade"))
            if not nif:
                continue
            self._by_nif.setdefault(nif, []).append(row)

    def _to_funding(self, raw: dict) -> PT2030Funding:
        return PT2030Funding(
            operation_code=str(raw.get("Código da Operação") or "").strip(),
            entity_name=str(raw.get("Designação da entidade") or "").strip(),
            role=str(raw.get("Papel da entidade") or "").strip(),
            beneficiary_percentage=_safe_float(
                raw.get("Percentagem Beneficiário na Operação")
            ),
            value_contractualized=_safe_float(raw.get("Valor contratualizado")),
            fund_approved=_safe_float(raw.get("Fundo Aprovado")),
            fund_executed=_safe_float(raw.get("Fundo Executado")),
            fund_paid=_safe_float(raw.get("Fundo Pago")),
            framework=str(raw.get("Enquadramento") or "").strip(),
        )

    async def search_by_nif(self, nif: str) -> dict[str, Any]:
        await self._ensure_loaded()
        nif = _normalize_nif(nif)

        raw = self._by_nif.get(nif, [])
        fundings = [self._to_funding(r) for r in raw]

        total_approved = sum(f.fund_approved or 0 for f in fundings)
        total_paid = sum(f.fund_paid or 0 for f in fundings)

        return {
            "pt2030_fundings": fundings,
            "has_pt2030_funding": bool(fundings),
            "pt2030_total_fund_approved": total_approved,
            "pt2030_total_fund_paid": total_paid,
        }
